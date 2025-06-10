import pandas as pd
import openpyxl  # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment  # type: ignore
from config import ARQUIVO_RESPOSTAS, ARQUIVO_SOCIOECONOMICO, ARQUIVO_EXCEL, NOME_DA_PLANILHA_RESULTADOS, NOME_DA_PLANILHA_SOCIO, COLUNA_NOME_SOCIO, NOME_DA_PLANILHA_MATERIAS
from analise_resultados import ler_mapeamento_materias

# --- Funções de Processamento ---

def pinta_materias(ws, mapeamento_materias, linha_cabecalho, start_column):
    """Cria e formata os cabeçalhos para os acertos de cada matéria."""
    
    # Estilos para os novos cabeçalhos (os mesmos dos cabeçalhos existentes)
    FILL_CABECALHO = PatternFill(start_color='568CB8', end_color='568CB8', fill_type='solid')
    FONTE_NEGRITO = Font(bold=True)
    ALINHAMENTO_CENTRO = Alignment(horizontal='center', vertical='center')
    
    current_col = start_column
    for materia in mapeamento_materias.keys():
        # Cabeçalho de Acertos por matéria
        acertos_cell = ws.cell(row=linha_cabecalho, column=current_col, value=f"Acertos {materia}")
        acertos_cell.fill = FILL_CABECALHO
        acertos_cell.font = FONTE_NEGRITO
        acertos_cell.alignment = ALINHAMENTO_CENTRO
        
        # Cabeçalho de % de Acertos por matéria
        percent_cell = ws.cell(row=linha_cabecalho, column=current_col + 1, value=f"% Acertos {materia}")
        percent_cell.fill = FILL_CABECALHO
        percent_cell.font = FONTE_NEGRITO
        percent_cell.alignment = ALINHAMENTO_CENTRO
        
        current_col += 2
    return current_col # Retorna o número total de colunas para o estilo geral

def aplicar_estilos_base(ws, linha_cabecalho, total_colunas_formatar, total_alunos):
    """Aplica a formatação estética básica (cores de cabeçalho e coluna de nome)."""
    print("Aplicando formatação estética...")
    
    FILL_CABECALHO = PatternFill(start_color='568CB8', end_color='568CB8', fill_type='solid')
    FILL_ALUNO_COL = PatternFill(start_color='90B3D0', end_color='90B3D0', fill_type='solid')
    FONTE_NEGRITO = Font(bold=True)
    
    # Estilo de todos os Cabeçalhos (Linha 5)
    for col in range(1, total_colunas_formatar + 1):
        ws.cell(row=linha_cabecalho, column=col).fill = FILL_CABECALHO
        ws.cell(row=linha_cabecalho, column=col).font = FONTE_NEGRITO

    # Estilo da Coluna de Nomes
    for linha in range(linha_cabecalho + 1, linha_cabecalho + total_alunos + 2):
        ws.cell(row=linha, column=1).fill = FILL_ALUNO_COL
        ws.cell(row=linha, column=1).font = FONTE_NEGRITO

def processar_provas(formatar_estilos=True):
    """
    Lê nomes de um arquivo, respostas de outro, e os combina na planilha final
    existente, calculando resultados gerais e por matéria.
    """
    # --- Passo 1: Carregar todos os dados de origem ---
    try:
        df_respostas = pd.read_csv(ARQUIVO_RESPOSTAS)
        df_socio = pd.read_excel(ARQUIVO_SOCIOECONOMICO, sheet_name=NOME_DA_PLANILHA_SOCIO)
    except FileNotFoundError as e:
        print(f"❌ ERRO: Arquivo de entrada não encontrado: {e.filename}.")
        return
    except ValueError as e:
        print(f"❌ ERRO: {e}. Verifique se a aba '{NOME_DA_PLANILHA_SOCIO}' existe no arquivo '{ARQUIVO_SOCIOECONOMICO}'.")
        return
    except KeyError as e:
        print(f"❌ ERRO: Coluna {e} não encontrada. Verifique os cabeçalhos nos arquivos de entrada.")
        return
    except PermissionError:
        print(f"❌ ERRO: Um arquivo está aberto em outro programa. Feche-o e tente novamente.")
        return

    # --- Passo 2: Preparar os dados ---
    gabarito = df_respostas.iloc[0]
    respostas_alunos = df_respostas.iloc[1:]
    nomes_ordenados = sorted(df_socio[COLUNA_NOME_SOCIO].dropna().unique())
    total_questoes = len(gabarito)

    # --- Passo 3: Carregar a planilha Excel EXISTENTE ---
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
        if NOME_DA_PLANILHA_RESULTADOS in workbook.sheetnames:
            ws = workbook[NOME_DA_PLANILHA_RESULTADOS]
        else:
            ws = workbook.create_sheet(NOME_DA_PLANILHA_RESULTADOS)
            print(f"Aviso: A planilha '{NOME_DA_PLANILHA_RESULTADOS}' não foi encontrada e foi criada.")
    except FileNotFoundError:
        print(f"❌ ERRO: A planilha final '{ARQUIVO_EXCEL}' não foi encontrada. Crie o arquivo ou verifique o nome.")
        return
    except PermissionError:
        print(f"❌ ERRO: O arquivo '{ARQUIVO_EXCEL}' está aberto. Feche-o e tente novamente.")
        return
    
    mapeamento_materias = ler_mapeamento_materias(ARQUIVO_EXCEL, NOME_DA_PLANILHA_MATERIAS)
    if not mapeamento_materias:
        print(f"❌ ERRO: Não foi possível ler o mapeamento de matérias.")
        return
    
    VERDE_CLARO = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    VERMELHO_CLARO = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # --- Passo 4: Escrever cabeçalhos ---

    linha_cabecalho = 5
    cabecalhos_questoes = df_respostas.columns

    ws.cell(row=linha_cabecalho, column=1, value="Aluno/Questões")
    for i, nome_coluna in enumerate(cabecalhos_questoes, start=2):
        numero_questao = nome_coluna.split('_')[-1]
        ws.cell(row=linha_cabecalho, column=i, value=int(numero_questao))
        
    ws.cell(row=linha_cabecalho, column=total_questoes + 2, value="Acertos")
    ws.cell(row=linha_cabecalho, column=total_questoes + 3, value="% de Acertos")
    
    linha_atual = linha_cabecalho + 1
    ws.cell(row=linha_atual, column=1, value="Gabarito")
    for i, questao in enumerate(cabecalhos_questoes, start=2):
        ws.cell(row=linha_atual, column=i, value=gabarito[questao])
    
    # Chama a função para criar os cabeçalhos das matérias
    coluna_final = pinta_materias(ws, mapeamento_materias, linha_cabecalho, start_column=total_questoes + 4)
    
    # --- Passo 5: Iterar e preencher dados ---
    num_max_linhas = max(len(nomes_ordenados), len(respostas_alunos))

    for i in range(num_max_linhas):
        linha_atual += 1

        # Preenche o nome do aluno na coluna 1
        if i < len(nomes_ordenados):
            ws.cell(row=linha_atual, column=1, value=nomes_ordenados[i])
        
        # Se houver uma linha de resposta correspondente, processa os dados
        if i < len(respostas_alunos):
            respostas_do_aluno = respostas_alunos.iloc[i]
            acertos_gerais = 0
            
            # Calcula acertos GERAIS
            for col_idx, questao_nome in enumerate(cabecalhos_questoes, start=2):
                resposta = respostas_do_aluno[questao_nome]
                celula_atual = ws.cell(row=linha_atual, column=col_idx, value=resposta)
                
                if str(resposta) == str(gabarito[questao_nome]):
                    acertos_gerais += 1
                    celula_atual.fill = VERDE_CLARO
                else:
                    celula_atual.fill = VERMELHO_CLARO

            # Escreve resultados GERAIS
            porcentagem_geral = (acertos_gerais / total_questoes) * 100
            ws.cell(row=linha_atual, column=total_questoes + 2, value=acertos_gerais)
            ws.cell(row=linha_atual, column=total_questoes + 3, value=porcentagem_geral).number_format = '0.00"%"'

            coluna_materia_atual = total_questoes + 4
            for materia, questoes_nums in mapeamento_materias.items():
                acertos_materia = 0
                for col_idx, questao_nome in enumerate(cabecalhos_questoes, start=2):
                    if int(questao_nome) in questoes_nums:
                        if str(respostas_do_aluno[questao_nome]) == str(gabarito[questao_nome]):
                            acertos_materia += 1

                total_questoes_materia = len(questoes_nums)
                percentual_materia = (acertos_materia / total_questoes_materia) * 100 if total_questoes_materia > 0 else 0

                ws.cell(row=linha_atual, column=coluna_materia_atual, value=acertos_materia)
                ws.cell(row=linha_atual, column=coluna_materia_atual + 1, value=percentual_materia).number_format = '0.00"%"'
                
                coluna_materia_atual += 2

    # --- Passo 6: Aplicar estilos e salvar ---
    if formatar_estilos:
        # Passa o número total de colunas para formatação correta dos cabeçalhos
        aplicar_estilos_base(ws, linha_cabecalho, coluna_final - 1, len(nomes_ordenados))
    
    try:
        workbook.save(ARQUIVO_EXCEL)
        print(f"✅ Processamento concluído! O arquivo '{ARQUIVO_EXCEL}' foi atualizado.")
    except PermissionError:
        print(f"❌ ERRO: Feche o arquivo '{ARQUIVO_EXCEL}' para poder salvá-lo.")

# --- Ponto de Entrada Principal ---
if __name__ == "__main__":

    processar_provas()